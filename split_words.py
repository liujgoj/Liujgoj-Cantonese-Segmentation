# 读取文件1.txt
with open('1.txt', 'r', encoding='UTF-8') as f:
    content = f.read()

# 将大写字母转为小写
content = content.lower()

# 替换字符串
# 將“可唔可以”替換爲“”可以唔可以”
content = content.replace("hor' mxhoryiq", "horyiq mxhoryiq")
# 將“可唔可能”替換爲“可能唔可能”
content = content.replace("hor' mxhornangx", "hornangx mxhornangx")
# 將縮寫“'哋”替換爲“佢哋”
content = content.replace("'deih", "keoiqdeih")
# 將“小完便”替換爲“小便”
content = content.replace("siur-yvnx-binh", "siurbinh")
# 將“接慣生”替換爲“接生”
content = content.replace("zip-gwaan-saangj", "zipsaangj")
# 將“bin-ganr”替換爲“bin”
content = content.replace("bin-ganr", "bin")
# 划分单词
words = content.split()

# 保存修改后的结果到文件2.txt
with open('2.txt', 'w', encoding='UTF-8') as f:
    f.write('\n'.join(words))


import re

# 读取文件2.txt
with open("2.txt", "r", encoding="utf-8") as file:
    content = file.read()

# 按照空格和连字符“-”以外所有的标点符号划分单词
words = re.findall(r'[\w\-]+', content)

# 对每个单词进行处理
for i in range(len(words)):
    word = words[i]

    # 去掉所有的数字，去掉除了连字符“-”以外所有的标点符号
    word = re.sub(r'\d', '', word)
    word = re.sub(r'[^\w\-]', '', word)

    # 删除特定的字符串(前綴同後綴)
    word = re.sub(r'^(aa-|louq-|daaih-|sai-|siur-|feix-)', '', word)
    word = re.sub(r'(-yix|-zeh|-soex|-foj|-saangj|-yij|-goj|-goh|-baak|-sukj|-hingj|-zungr|-taair|-zair|-neoiq|-neoir|-maaj|-noengr|-gaaj|-gei|-sih|-yex|-fur|-zer|-zej|-gungj|-ciux|-satj|-wongx|-guj|-samr|-zikh|-siu)$', '', word)

    # 处理特殊情况
    # 將“冇事冇幹”等替換爲“冇 事幹”
    if re.match(r'^(mouq)-([a-z]+)-(mouq)-([a-z]+)$', word):
        word = re.sub(r'^(mouq)-', 'mouq ', word)
        word = re.sub(r'-(mouq)-', '', word)
    # 將“有忠有義”等替換爲“有 忠義”
    elif re.match(r'^(yauq)-([a-z]+)-(yauq)-([a-z]+)$', word):
        word = re.sub(r'^(yauq)-', 'yauq ', word)
        word = re.sub(r'-(yauq)-', '', word)
    # 將單詞中間嘅“ganr”等時態同埋中間插入嘅單詞刪除
    elif re.match(r'^[a-z]+(ganr|zor|zvh|gwo)(-[a-z]+)+-[a-z]+$', word):
        word = re.sub(r'(ganr|zor|zvh|gwo)(-[a-z]+)+-', '', word)
    # 將單詞中間嘅“ganr”等時態刪除
    elif re.match(r'^[a-z]+(ganr|zor|zvh|gwo)-[a-z]+$', word):
        word = re.sub(r'(ganr|zor|zvh|gwo)-', '', word)
    # 將插入單詞中間嘅音節同埋單詞刪除
    elif re.match(r'^[a-z]+(-[a-z]+)+-[a-z]+$', word) and word.count('-') >= 2:
        word = re.sub(r'(-[a-z]+)+-', '', word)

    # 将处理后的单词保存到列表中
    words[i] = word.strip()

# 将处理后的单词写入文件3.txt，每行一个单词
with open("3.txt", "w", encoding="utf-8") as file:
    for word in words:
        file.write(word + "\n")

import openpyxl
import re

# 读取aspect.txt文件，获取其中的单词列表
with open("aspect.txt", "r", encoding="utf-8") as file:
    aspect_words = file.read().split()

# 读取文件3.txt
with open("3.txt", "r", encoding="utf-8") as file:
    content = file.read()

# 按照空格划分单词
words = content.split()

# 处理单词
for i in range(len(words)):
    word = words[i]

    # 检查单词是否在aspect.txt文件中
    if word in aspect_words:
        continue

    # 根据匹配规则处理单词
    elif re.match(r'^[a-z]+ganr$', word):
        word = re.sub(r'ganr', '', word)
    elif re.match(r'^[a-z]+zor$', word):
        word = re.sub(r'zor', '', word)
    elif re.match(r'^[a-z]+zvh$', word):
        word = re.sub(r'zvh', '', word)
    elif re.match(r'^[a-z]+gwo$', word):
        word = re.sub(r'gwo', '', word)

    # 将处理后的单词保存回列表中
    words[i] = word

# 统计单词数量和频率
word_counts = {}
for word in words:
    if word not in aspect_words:
        if word in word_counts:
            word_counts[word] += 1
        else:
            word_counts[word] = 1

total_words = len(words)
word_freqs = {word: count / total_words for word, count in word_counts.items()}

# 按照词频和词频数量排序
sorted_word_freqs = sorted(word_freqs.items(), key=lambda x: (x[1], word_counts[x[0]]), reverse=True)

# 将结果写入Excel文件中
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Word Frequencies"

# 写入表头
worksheet.cell(row=1, column=1, value="Word")
worksheet.cell(row=1, column=2, value="Count")
worksheet.cell(row=1, column=3, value="Frequency")

# 写入数据
for i, (word, freq) in enumerate(sorted_word_freqs):
    worksheet.cell(row=i+2, column=1, value=word)
    worksheet.cell(row=i+2, column=2, value=word_counts[word])
    worksheet.cell(row=i+2, column=3, value=freq)

# 写入总单词数量
worksheet.cell(row=len(sorted_word_freqs)+2, column=1, value="Total Words")
worksheet.cell(row=len(sorted_word_freqs)+2, column=2, value=total_words)

# 保存Excel文件
workbook.save(filename="4.xlsx")
